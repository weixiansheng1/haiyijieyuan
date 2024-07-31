"""
# =============================================================================
#                      工程维修联系单
# =============================================================================
"""
import time
import re
import os
from   PyQt5.QtWidgets import QMessageBox
from   PyQt5.QtSql   import  QSqlQuery
from   dataclasses   import dataclass
from   collections   import defaultdict

from   openpyxl        import Workbook
from   openpyxl.styles import Font, Border,Side,PatternFill,Alignment
from   openpyxl.worksheet.datavalidation import DataValidation

# 获取当前日期
now  = time.strftime("%Y-%m-%d"      ,time.localtime(time.time()))
now1 = time.strftime("%Y年 %m月 %d日",time.localtime(time.time()))
now2 = time.strftime("%Y%m"         ,time.localtime(time.time()))
now3 = time.strftime("%Y年%m月份"    ,time.localtime(time.time()))
now4 = time.strftime("%Y年%m月"      ,time.localtime(time.time())).replace('年0','年')
now5 = time.strftime("%Y年%m月%d日"  ,time.localtime(time.time())).replace('年0','年').replace('月0','月')
now6  = time.strftime("%Y.%m.%d"    ,time.localtime(time.time()))

@dataclass
class Write_engineering_repair_sheet:
    output_dir      :str          # 输出文件路径
    select_row      :None         # 选择好的行
    
    # ======================================================================= #
    #                              功能                                    
    # ======================================================================= #

    # 连接数据库并加载数据
    def connect_jieyuan_database(self):
                        #【   隐患内容  ,     隐患整改措施   ,     责任人员     ,    整改期限       】
        select_column = 'hidden_content,corrective_measures,responsible_punish,corrective_deadline'

        if len(self.select_row) ==1:
            select_row_ = self.select_row[0]
            exec_ = "SELECT {} FROM hidden WHERE Id = {} ".format(select_column,select_row_) 
        elif len(self.select_row)>=1:
            select_row_ = tuple(self.select_row)
            exec_ = "SELECT {} FROM hidden WHERE Id IN {} ".format(select_column,select_row_) 

        # print(exec_) 
        query = QSqlQuery(exec_)

        # 初始输入数据字符串
        str_hidden_content         = []    # 隐患内容   
        str_corrective_measures    = []    # 隐患整改措施
        responsible_person         = []    # 责任人员
        retification_deadline      = []    # 整改期限
        
        cache = [str_hidden_content,str_corrective_measures,responsible_person,retification_deadline]
        
        # 跟据query获取数据库的数据，并生成字符串
        while query.next():
            [cache[x].append(query.value(x)) for x in range(len(cache))]


        self.problem              = str_hidden_content                                # 存在问题 list
        self.retification_requirements =  str_corrective_measures                     # 整改要求
        self.responsible_person   = responsible_person                                # 责任人员
        self.retification_deadline= retification_deadline                             # 整改期限  
        
        ########## 数据清洗合并 ##########
        self.retification_requirements_and_deadline =  None # 整改要求与整改期限合并
        self.responsible_person_tostr               =  None # 责任人员

        # 数据清洗合并执行框
        self.merge_check_str()
        
        
    def function(self):
        self.connect_jieyuan_database()
        wb = Workbook()     # 创建一个 workbook
        ws = wb.active      # 获取被激活的 worksheet
        
        align = Alignment( horizontal='center', vertical = 'center', wrap_text = True ) # 居中类
        font_normul  = Font( name   = "宋体", size = 12 , bold = False , italic = False , color = "00000000")
        font_title   = Font( name   = "宋体", size = 18 , bold = True  , italic = False , color = "00000000")
        # font_number  = Font( name   = 'Times New Roman ' ,size = 12 , bold = False  , italic = False , color = "00000000")
        
        border_cell  = Border(left  = Side(style = 'thin', color = '00000000'),
                              right = Side(style = 'thin', color = '00000000'),
                              top   = Side(style = 'thin', color = "00000000"),
                              bottom= Side(style = 'thin', color = "00000000")
                              )
        fill_ =  PatternFill(patternType='solid',fgColor='ffe2efda')#淡绿色
        
        ################## 调整表头 ####################
        
        ws['A1'] = "工程维修联系单"                                   # 设置单元格内容
        ws['A1'].alignment = align                                   # 设置居中
        ws['A1'].font      = font_title                              # 设置字体
        ws['A1'].border    = border_cell                             # 设置边框
        
        # for row in ws['A3:J3']:
        #     for cell in row:
        #         cell.border    = border_cell                       # 设置边框
                
        ws.merge_cells(range_string = 'A1:D1')                       # 合并单元格
        ws.merge_cells(range_string = 'A4:A5')
        ws.merge_cells(range_string = 'B4:D4')
        ws.merge_cells(range_string = 'B5:D5')
        ws.merge_cells(range_string = 'C6:D6')
        
        ws['A2'] = '主送单位'
        ws['A3'] = '抄送单位'
        ws['A4'] = '项目名称'
        ws['A6'] = '序号'
        ws['B6'] = '存在问题'
        ws['C2'] = '编   号'
        ws['C3'] = '发文日期'
        ws['C6'] = '图片'

        # 设置表头各列宽度
        column_width = {'A':12.43,'B':44.22,'C':12.26,'D':41.43}
        for column_name,column_width in column_width.items():
            ws.column_dimensions[column_name].width = column_width 
            
        # 设置表头各行行高
        line_hight = [48,28,27,27,38,28]
        for i,l_hight in enumerate(line_hight,start=1):
            ws.row_dimensions[i].height = l_hight    
            
        # 设置字体
        for row in ws['A2:D6']:
            for cell in row:
                cell.alignment = align              # 设置居中
                cell.font      = font_normul        # 设置字体
                cell.border    = border_cell        # 设置边框

        ################## 填入数据 ####################
        ws['B2'] = 'ces'# self.responsible_person_tostr    # 主送单位
        ws['D2'] = f'{now2}-01'                     # 编号 
        # ws['D2'].font =  font_number   
        ws['B3'] = '请填写'                         # 抄送单位 
        ws['D3'] = now5                             # 发文日期
            
        ws['B4'] = '珠海中信生态环保产业园餐厨垃圾处理一期工程'  # 项目名称
        def set_upslide(self,ws, cell_alignment, upslid_list, allow_blank=True):
            """
            单元格设置下拉菜单
            :param ws: sheet对象
            :param cell_alignment: 单元格的范围，例如"G2:G10"
            :param upslid_list: 下拉选项，例如：'"确认漏洞,确认误报,确认报备,待确认"'
            :param allow_blank: 是否可以为空，默认为True，可以为空
            :return:
            """
            dv = DataValidation(type="list", formula1=upslid_list, allow_blank=allow_blank)
            dv.error = "Your entry is not in the list"
            dv.errorTitle = "Invalid Entry"
            dv.prompt = "Please select from the list"
            dv.promptTitle = "List Selection"
            dv.add(cell_alignment)
            ws.add_data_validation(dv)

        
        # # 批量合并A到J的类似A2:A3单元格，并填入字符串，并调整格式
        # a = [ 'A'  ,  'B'    ,   'C'    ,   'D'   ,        'E'     ,      'F'    ,  'J' ]
        # b = ['序号','检查时间','存在问题','整改要求','责任人/责任单位','海宜扣分情况','备注']
        # c = [8.01  ,  26.83  ,   45    ,  45      ,      32.1      ,     27.26   , 20.51]
        
        # for i, x in enumerate(a):
        #     ai = x +'2:%s3'%x
        #     bi = '%s2'%x
        #     ws.merge_cells(range_string = ai)
        #     ws[bi] = b[i]                         # 填入数据
        #     ws.column_dimensions[x].width = c[i]  # 调整列宽
        #     ws[bi].alignment = align              # 设置居中
        #     ws[bi].font      = font_normul        # 设置字体
        #     ws[bi].border    = border_cell        # 设置边框
            
        # ws.column_dimensions['G'].width = 20      # 调整列宽
        # ws.column_dimensions['H'].width = 20  
        # ws.column_dimensions['I'].width = 20  
        
        # ws.row_dimensions[1].height     = 54      # 调整行高
        # ws.row_dimensions[2].height     = 36
        # ws.row_dimensions[3].height     = 36

        # ws['G2'] = '内部扣分情况'                  # 填入数据
        # ws['G3'] = '责任人'
        # ws['H3'] = '责任管理人员'
        # ws['I3'] = '责任部门'
        
        # ws['G2'].alignment = align                # 设置居中
        # ws['G3'].alignment = align
        # ws['H3'].alignment = align
        # ws['I3'].alignment = align
        
        ###################### 添加内容 ###################
        self.connect_jieyuan_database()           # 加载数据

        
       #  count     = 0
       #  row_start = 3
       #  for i, classify in enumerate(append_wb):
       #      emerge_name = append_name[i]
       #      print('classify',classify)
       #      if classify:
       #          count = row_start   
       #          for check_time, data in classify.items():

       #              for b in data:
       #                  row_start += 1
       #                  ws.append([row_start-3, 
       #                             check_time,
       #                             self.problem[b],
       #                             self.retification_requirements_and_deadline[b], 
       #                             self.responsible_person_and_unit[b],
       #                             '-',
       #                             self.responsible_person_tostr  [b],             # 责任人员
       #                             self.responsible_person_g_tostr[b],             # 责任管理人员
       #                             self.responsible_person_m_tostr[b],             # 责任部门
       #                             emerge_name,
       #                             ])  # 设置一行内容
                        
       # # 调整格式
       #                  ws.row_dimensions[row_start].height     = 179              # 调整行高
       #              ws.merge_cells(range_string = 'B%s:B%s'%(row_start-len(data)+1,row_start)) # 合并 检查时间 列
                    
       #          if append_name[i] == '海宜检查':                                    # 单元格填充
       #              for row in ws['A4:J%s'%row_start]:
       #                  for cell in row:
       #                      cell.fill = fill_
       #          ws.merge_cells(range_string = 'J%s:J%s'%(count+1,row_start))                   # 合并 备注 列 

       #  for row in ws['A4:J%s'%row_start]:

       #      for cell in row:
       #          cell.alignment = align              # 设置居中
       #          cell.font      = font_normul        # 设置字体
       #          cell.border    = border_cell


        output_dir = r"{}".format(self.output_dir)
        self.output = os.path.join(output_dir, f"工程维修联系单（餐厨项目xxx隐患）-{now6}.xlsx")
        wb.save(self.output)  
    
    def output_file_path(self):
        return  self.output

    #   数据清洗合并 
    def merge_check_str(self): 
        
        ################# 计算 #################
        retification_requirements_and_deadline = [] # 整改要求与整改期限合并
        responsible_person_tostr               = [] # 责任人
        responsible_person_and_unit            = []
        person_score = '-'

        for i in range(len(self.retification_requirements)):
            str_right_deadlint = ''
            # 整理整改期限
            if self.retification_deadline[i] != '':

                if '即时整改' in self.retification_deadline[i] :
                    ...
                elif '-' in self.retification_deadline[i] :
                    str_right_deadlint = '前整改完成'
                    self.retification_deadline[i] = self.retification_deadline[i].replace('-', '年', 1).replace('-', '月')+'日' # str.replace(old, new[, max])
                    self.retification_deadline[i] =  self.retification_deadline[i].replace('年0','年').replace('月0','月')
            else:
                self.retification_deadline[i] = '警告！请填写整改期限' 
                self.show_warning_message('警告！请填写整改期限')

            retification_requirements_and_deadline.append(self.retification_requirements[i] + '（整改期限：%s%s）'%(self.retification_deadline[i],str_right_deadlint  ))


            # 责任人员
            if self.responsible_person[i]!= '':
                if '公司' in  self.responsible_person[i]:
                    str_front = '责任单位：'
                elif ('部'or '中心') in  self.responsible_person[i]:
                    str_front = '责任部门：'
                else:
                    str_front = '责任人员：'
                    if '-' in self.responsible_person[i]:
                        person_score = re.sub(r'(（\w+）)?-(\d+(?:\.\d+)?)', r'\1扣\2分', self.responsible_person[i])
                    else:
                        person_score = '-'
            else:
                str_front,person_score = ''
                self.responsible_person[i] = ''

            person = self.responsible_person[i].translate(str.maketrans('-', ' ','0123456789'))  # translate(str.maketrans: 第一个和第二个参数的长度必须匹配。在两个参数的情况下，会将第一个参数的字符，依次的映射成第二个参数的字符（o-> X，w-> Y）。第三个参数表示在映射完的结果之后，需要移除的字符。
            responsible_person_and_unit.append( str_front + person )
            responsible_person_tostr.append(person_score )
            
        

        # 合并结果
        self.retification_requirements_and_deadline =  retification_requirements_and_deadline # 整改要求与整改期限合并
        
        self.responsible_person_tostr               =  responsible_person_tostr               # 责任人
        print(self.retification_requirements_and_deadline)
   
    #%%    显示错误提示信息
    def show_warning_message(self,show_error_str):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Warning)
        message_box.setText("{}".format(show_error_str))
        message_box.setWindowTitle("警告")
        message_box.exec_()

# =============================================================================
#           测试
# =============================================================================
# if  __name__ == '__main__':

#     select_data = [0,1,2,3,4,5,6]
#     taxt = Write_engineering_repair_sheet("../main",select_data)
#     taxt.function() # 主函数
