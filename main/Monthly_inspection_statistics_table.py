"""
# =============================================================================
#                      每月检查情况统计表
# =============================================================================
"""
import time
import re
import os
from   PyQt5.QtWidgets import QMessageBox
from   PyQt5.QtSql   import  QSqlQuery
from   dataclasses   import dataclass
from   collections   import defaultdict

from   openpyxl        import Workbook
from   openpyxl.styles import Font, Border,Side,PatternFill,Alignment

# 获取当前日期
now  = time.strftime("%Y-%m-%d"      ,time.localtime(time.time()))
now1 = time.strftime("%Y年 %m月 %d日",time.localtime(time.time()))
now2 = time.strftime("%Y0%m"         ,time.localtime(time.time()))
now3 = time.strftime("%Y年%m月份"    ,time.localtime(time.time()))
now4 = time.strftime("%Y年%m月"      ,time.localtime(time.time())).replace('年0','年')
now5 = time.strftime("%Y年%m月%d日"  ,time.localtime(time.time())).replace('年0','年').replace('月0','月')

@dataclass
class Write_inspect_the_situation_sheet:
    output_dir      :str          # 输出文件路径
    select_row      :None         # 选择好的行
    
    # ======================================================================= #
    #                              功能                                    
    # ======================================================================= #

    # 连接数据库并加载数据
    def connect_jieyuan_database(self):
                     #【       被检查单位     ,  检查类型, 检查时间  ,   隐患内容   ,     隐患整改措施   ,     责任人员      ,    责任管理人员           ，     责任部门               ，  整改期限       】
        select_column = 'unit_to_be_inspected,check_type,check_time,hidden_content,corrective_measures,responsible_punish,responsible_manager_punish,responsible_department_punish,corrective_deadline'

        if len(self.select_row) ==1:
            select_row_ = self.select_row[0]
            exec_ = "SELECT {} FROM hidden WHERE Id = {} ".format(select_column,select_row_) 
        elif len(self.select_row)>=1:
            select_row_ = tuple(self.select_row)
            exec_ = "SELECT {} FROM hidden WHERE Id IN {} ".format(select_column,select_row_) 

        #print(exec_) 
        query = QSqlQuery(exec_)

        # 初始输入数据字符串
        str_unit_to_be_inspected   = []    # 被检查单位 
        str_check_type             = []    # 检查类型
        str_check_time             = []    # 检查时间
        str_hidden_content         = []    # 隐患内容   
        str_corrective_measures    = []    # 隐患整改措施
        responsible_person         = []    # 责任人员
        responsible_manager_punish = []    # 责任管理人员
        responsible_department_punish = [] # 责任部门
        retification_deadline      = []    # 整改期限
        
        cache = [str_unit_to_be_inspected,str_check_type,str_check_time,str_hidden_content,str_corrective_measures,responsible_person,responsible_manager_punish,responsible_department_punish,retification_deadline]
        
        # 跟据query获取数据库的数据，并生成字符串
        while query.next():
            [cache[x].append(query.value(x)) for x in range(len(cache))]

        # 检查时间格式转换
        to_time  = [str_check_time[i].replace('-', '年', 1).replace('-', '月').replace('年0','年').replace('月0','月')+'日' for i in range(len(str_check_time))]

        self.check_time           = to_time                                           # 检查时间   
        self.str_check_type       = str_check_type                                    # 检查类型   use
        self.unit_to_be_inspected = str_unit_to_be_inspected                          # 被检查单位
        self.problem              = str_hidden_content                                # 存在问题   use
        self.retification_requirements =  str_corrective_measures                     # 整改要求
        self.responsible_person   = responsible_person                                # 责任人员
        self.responsible_person_g = responsible_manager_punish                        # 责任管理人员
        self.responsible_person_m = responsible_department_punish                     # 责任部门
        self.retification_deadline= retification_deadline                             # 整改期限  
        
        ########## 数据清洗合并 ##########
        self.retification_requirements_and_deadline =  None # 整改要求与整改期限合并
        self.responsible_person_and_unit            =  None # 责任人/责任单位
        self.responsible_person_tostr               =  None # 责任人员
        self.responsible_person_g_tostr             =  None # 责任管理人员
        self.responsible_person_m_tostr             =  None # 责任部门
        
        # 数据清洗合并执行框
        self.merge_check_str()
        
        
    def function(self):

        wb = Workbook()     # 创建一个 workbook
        ws = wb.active      # 获取被激活的 worksheet
        
        align = Alignment( horizontal='center', vertical = 'center', wrap_text = True ) # 居中类
        font_normul  = Font( name   = "宋体", size = 14 , bold = False , italic = False , color = "00000000")
        font_title   = Font( name   = "宋体", size = 20 , bold = True  , italic = False , color = "00000000")
        border_cell  = Border(left  = Side(style = 'thin', color = '00000000'),
                              right = Side(style = 'thin', color = '00000000'),
                              top   = Side(style = 'thin', color = "00000000"),
                              bottom= Side(style = 'thin', color = "00000000")
                              )
        fill_ =  PatternFill(patternType='solid',fgColor='ffe2efda')#淡绿色
        
        ################## 调整表头 ####################
        
        ws['A1'] = "洁源公司%s安全生产专项工作检查情况统计表"%now4      # 设置单元格内容
        ws['A1'].alignment = align                                   # 设置居中
        ws['A1'].font      = font_title                              # 设置字体
        ws['A1'].border    = border_cell                             # 设置边框
        
        for row in ws['A3:J3']:
            for cell in row:
                cell.border    = border_cell                         # 设置边框
                
        ws.merge_cells(range_string = 'A1:J1')                       # 合并单元格
        ws.merge_cells(range_string = 'G2:I2')

        # 批量合并A到J的类似A2:A3单元格，并填入字符串，并调整格式
        a = [ 'A'  ,  'B'    ,   'C'    ,   'D'   ,        'E'     ,      'F'    ,  'J' ]
        b = ['序号','检查时间','存在问题','整改要求','责任人/责任单位','海宜扣分情况','备注']
        c = [8.01  ,  26.83  ,   45    ,  45      ,      32.1      ,     27.26   , 20.51]
        
        for i, x in enumerate(a):
            ai = x +'2:%s3'%x
            bi = '%s2'%x
            ws.merge_cells(range_string = ai)
            ws[bi] = b[i]                         # 填入数据
            ws.column_dimensions[x].width = c[i]  # 调整列宽
            ws[bi].alignment = align              # 设置居中
            ws[bi].font      = font_normul        # 设置字体
            ws[bi].border    = border_cell        # 设置边框
            
        ws.column_dimensions['G'].width = 20      # 调整列宽
        ws.column_dimensions['H'].width = 20  
        ws.column_dimensions['I'].width = 20  
        
        ws.row_dimensions[1].height     = 54      # 调整行高
        ws.row_dimensions[2].height     = 36
        ws.row_dimensions[3].height     = 36

        ws['G2'] = '内部扣分情况'                  # 填入数据
        ws['G3'] = '责任人'
        ws['H3'] = '责任管理人员'
        ws['I3'] = '责任部门'
        
        ws['G2'].alignment = align                # 设置居中
        ws['G3'].alignment = align
        ws['H3'].alignment = align
        ws['I3'].alignment = align
        
        ###################### 添加内容 ###################
        self.connect_jieyuan_database()           # 加载数据
        # 分类
        self.check_haiyi_dict       = self.classify_hz('海宜查',self.str_check_type,self.check_time)        # 检查类型:海宜查
        self.check_jieyuan_dickt    = self.classify_hz('自查'  ,self.str_check_type,self.check_time)        # 检查类型:自查
        self.check_shouyun_dict     = self.classify_hz('查收运'  ,self.str_check_type,self.check_time)      # 检查类型:查收运
        self.check_emergency_dict   = self.classify_hz('查行政与应急'  ,self.str_check_type,self.check_time)# 检查类型:查行政与应急
        
        append_name = ['海宜检查'            ,'洁源自查'              ,'洁源自查收运'          ,'查行政与应急'            ]
        append_wb   = [self.check_haiyi_dict,self.check_jieyuan_dickt,self.check_shouyun_dict,self.check_emergency_dict]
        
        count     = 0
        row_start = 3
        for i, classify in enumerate(append_wb):
            emerge_name = append_name[i]
            print('classify',classify)
            if classify:
                count = row_start   
                for check_time, data in classify.items():

                    for b in data:
                        row_start += 1
                        ws.append([row_start-3, 
                                   check_time,
                                   self.problem[b],
                                   self.retification_requirements_and_deadline[b], 
                                   self.responsible_person_and_unit[b],
                                   '-',
                                   self.responsible_person_tostr  [b],             # 责任人员
                                   self.responsible_person_g_tostr[b],             # 责任管理人员
                                   self.responsible_person_m_tostr[b],             # 责任部门
                                   emerge_name,
                                   ])  # 设置一行内容
                        
       # 调整格式
                        ws.row_dimensions[row_start].height     = 179              # 调整行高
                    ws.merge_cells(range_string = 'B%s:B%s'%(row_start-len(data)+1,row_start)) # 合并 检查时间 列
                    
                if append_name[i] == '海宜检查':                                    # 单元格填充
                    for row in ws['A4:J%s'%row_start]:
                        for cell in row:
                            cell.fill = fill_
                ws.merge_cells(range_string = 'J%s:J%s'%(count+1,row_start))                   # 合并 备注 列 

        for row in ws['A4:J%s'%row_start]:

            for cell in row:
                cell.alignment = align              # 设置居中
                cell.font      = font_normul        # 设置字体
                cell.border    = border_cell

        output_dir = r"{}".format(self.output_dir)
        self.output = os.path.join(output_dir, "洁源公司%s安全生产专项工作检查情况统计表.xlsx"%now4)
        wb.save(self.output)  # 保存 Excel 文件 #TODO:输出文件位置待测试
    
    def output_file_path(self):
        return  self.output

    # 预分类器：用于分类输入数据 返回分类好的字典
    def Pre_sorted(self,pre):
        classified_dates_dict  = defaultdict(list)     # 创建一个默认字典来存储日期分类及对应的序号列表
        for index, data in enumerate(pre):             # 遍历日期列表，将日期及其对应的序号存入字典中
            classified_dates_dict[data].append(index)
        return classified_dates_dict
    
    # 二次分类器：输入要的类型，返回该类型所对应的日期的分类字典，字典里包含有分类好的列表
    def classify_hz(self,object_to_c:str,object_list,Sec_class_ob): # 待分类对象（str），对象的列表,二次分类对象
        cache = []
        object_newdict = {}
        sec_newdict    = {}
        date = self.Pre_sorted(object_list)
        print(date)
        for data, index_list in date.items():
            # print(data, ":", index_list)
            if object_to_c in  data:                     # 选出对象列表
                object_newdict[object_to_c] = index_list # 字典
                
        print('第一次分类情况：',object_newdict)
        if object_to_c in object_newdict.keys():
            # 二次分类
            data = [cache.append(Sec_class_ob[i]) for i in object_newdict[object_to_c]]
            # print('第二次预分类:',cache)
            Sec_date = self.Pre_sorted(cache)
            # print('第二次分类情况：',Sec_date )
            for data, index_list in Sec_date.items():
                sec_newdict[data] = [object_newdict[object_to_c][i]  for i in index_list ] # 将列表对应回去
                
            print('最终分类：',sec_newdict)
            
            return sec_newdict
        else:
            return None


    #   数据清洗合并 
    def merge_check_str(self): 
        
        ################# 计算 #################
        retification_requirements_and_deadline = [] # 整改要求与整改期限合并
        responsible_person_and_unit            = [] # 责任人/责任单位
        responsible_person_tostr               = [] # 责任人
        responsible_person_g_tostr             = [] # 责任管理人员
        responsible_person_m_tostr             = [] # 责任部门
        
        person_score = '-'

        for i in range(len(self.retification_requirements)):
            str_right_deadlint = ''
            # 整理整改期限
            if self.retification_deadline[i] != '':

                if '即时整改' in self.retification_deadline[i] :
                    ...
                elif '-' in self.retification_deadline[i] :
                    str_right_deadlint = '前整改完成'
                    self.retification_deadline[i] = self.retification_deadline[i].replace('-', '年', 1).replace('-', '月')+'日' # str.replace(old, new[, max])
            else:
                self.retification_deadline[i] = '警告！请填写整改期限' 
                self.show_warning_message('警告！请填写整改期限')

            retification_requirements_and_deadline.append(self.retification_requirements[i] + '\n（整改期限：%s%s）'%(self.retification_deadline[i],str_right_deadlint  ))

            # 替换责任管理人员扣分
            if self.responsible_person_g[i] != '':
                g =  re.sub(r'(（\w+）)?-(\d+(?:\.\d+)?)', r'\1扣\2分', self.responsible_person_g[i])
            else:
                g = '-'
            responsible_person_g_tostr.append(g)

            # 替换责任部门扣分
            if self.responsible_person_m[i] != '':
                m =  re.sub(r'(（\w+）)?-(\d+(?:\.\d+)?)', r'\1扣\2分', self.responsible_person_m[i])
            else:
                m = '-'
            responsible_person_m_tostr.append(m)

            # 责任人员
            if self.responsible_person[i]!= '':
                if '公司' in  self.responsible_person[i]:
                    str_front = '责任单位：'
                elif ('部'or '中心') in  self.responsible_person[i]:
                    str_front = '责任部门：'
                else:
                    str_front = '责任人员：'
                    if '-' in self.responsible_person[i]:
                        person_score = re.sub(r'(（\w+）)?-(\d+(?:\.\d+)?)', r'\1扣\2分', self.responsible_person[i])
                    else:
                        person_score = '-'
            else:
                str_front,person_score = ''
                self.responsible_person[i] = ''

            person = self.responsible_person[i].translate(str.maketrans('-', ' ','0123456789'))  # translate(str.maketrans: 第一个和第二个参数的长度必须匹配。在两个参数的情况下，会将第一个参数的字符，依次的映射成第二个参数的字符（o-> X，w-> Y）。第三个参数表示在映射完的结果之后，需要移除的字符。
            responsible_person_and_unit.append( str_front + person )
            responsible_person_tostr.append(person_score )
        # print(responsible_person_m_tostr)

        # 合并结果
        self.retification_requirements_and_deadline =  retification_requirements_and_deadline # 整改要求与整改期限合并
        self.responsible_person_and_unit            =  responsible_person_and_unit            # 责任人/责任单位
        self.responsible_person_tostr               =  responsible_person_tostr               # 责任人
        self.responsible_person_g_tostr             =  responsible_person_g_tostr             # 责任管理人员
        self.responsible_person_m_tostr             =  responsible_person_m_tostr             # 责任部门
   
    #%%    显示错误提示信息
    def show_warning_message(self,show_error_str):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Warning)
        message_box.setText("{}".format(show_error_str))
        message_box.setWindowTitle("警告")
        message_box.exec_()

# =============================================================================
#           测试
# =============================================================================
# select_data = [0,1,2,3,4,5,6]

# taxt = Write_inspect_the_situation("../main",select_data)
# taxt.function() # 主函数
