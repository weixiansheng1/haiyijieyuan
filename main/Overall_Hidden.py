# -*- coding: utf-8 -*-
"""
# =============================================================================
#                      隐患总台账
# =============================================================================
"""

import time
import os
from   PyQt5.QtSql           import QSqlQuery
from   PyQt5.QtWidgets       import QMessageBox
from   dataclasses           import dataclass
from   collections           import defaultdict
from   openpyxl              import Workbook
from   openpyxl.styles       import Font, Border,Side,PatternFill,Alignment
from   openpyxl.chart        import PieChart,Reference
from   openpyxl.chart.label  import DataLabelList
from   openpyxl.chart.layout import Layout, ManualLayout



# 获取当前日期
now  = time.strftime("%Y-%m-%d"      ,time.localtime(time.time()))
now1 = time.strftime("%Y年 %m月 %d日",time.localtime(time.time()))
now2 = time.strftime("%Y0%m"         ,time.localtime(time.time()))
now3 = time.strftime("%Y年%m月份"    ,time.localtime(time.time()))
now4 = time.strftime("%Y年%m月"      ,time.localtime(time.time())).replace('年0','年')
now5 = time.strftime("%Y年%m月%d日"  ,time.localtime(time.time())).replace('年0','年').replace('月0','月')

@dataclass
class overall_hidden_sheet():
    output_dir   :str          # 输出文件路径
    select_row   :None
    ############## 数据处理 ##############
    
    # 删除列表中空字符
    def not_empty(self,s):
        return s and s.strip()
    # 连接数据库
    def connect_jieyuan_database(self):
                     #【       0       ,      1            ,         2         ，    3      ,     4     ,     5    ,          6           ,           7               ,         8        ,             9                  ,    10 】
        select_column = 'hidden_content,corrective_measures,corrective_deadline,hidden_type,hidden_level,check_time,correactive_situastion,retification_complete_time,responsible_punish,retification_confirmation_number,check_type' 

        if len(self.select_row) ==1:
            select_row_ = self.select_row[0]
            exec_ = "SELECT {} FROM hidden WHERE Id = {} ".format(select_column,select_row_) 
        elif len(self.select_row)>=1:
            select_row_ = tuple(self.select_row)
            exec_ = "SELECT {} FROM hidden WHERE Id IN {} ".format(select_column,select_row_) 

        #print(exec_) 
        query = QSqlQuery(exec_)

        # 初始输入数据字符串
        str_hidden_content       = []    # 隐患内容
        str_corrective_measures  = []    # 隐患整改措施
        str_corrective_deadline  = []    # 隐患整改期限
        str_hidden_type          = []    # 隐患类型
        str_hidden_level         = []    # 隐患级别
        str_check_time           = []    # 检查时间
        str_correactive_situastion = []  # 整改情况
        str_complete_time        = []    # 整改完成时间
        str_responsible_punish   = []    # 责任人
        str_retification_confirmation_number = [] # 整改确认单编号
        str_check_type               = []    # 检查类型 ：海宜查、自查
        
        cache = [str_hidden_content,str_corrective_measures,str_corrective_deadline,str_hidden_type,str_hidden_level,str_check_time,str_correactive_situastion,str_complete_time,str_responsible_punish,str_retification_confirmation_number,str_check_type]
        
        # 跟据query获取数据库的数据，并生成字符串
        while query.next():
            [cache[x].append(query.value(x)) for x in range(len(cache))]
            pass

        # 检查时间格式转换
        to_time          = [str_check_time[i].replace('-', '年', 1).replace('-', '月').replace('年0','年').replace('月0','月')+'日' for i in range(len(str_check_time)) ]
        to_complete_time = [str_complete_time [i].replace('-', '年', 1).replace('-', '月').replace('年0','年').replace('月0','月')+'日'  if str_complete_time[i] else ''  for i in range(len(str_complete_time)) ]


###############################################################################################################
        self.str_hidden_content      = str_hidden_content                             # 隐患内容   
        self.str_corrective_measures = str_corrective_measures                                 # 隐患整改措施   
        
        self.str_corrective_deadline = str_corrective_deadline                        # 整改期限   
        self.str_hidden_type         =  str_hidden_type                               # 隐患类型
        self.str_hidden_level        = str_hidden_level                               # 隐患级别
        self.str_check_time          = to_time                                        # 检查时间
        self.str_correactive_situastion = str_correactive_situastion                  # 整改情况
        self.str_complete_time       = to_complete_time                              # 整改完成时间  
        self.str_responsible_punish  = str_responsible_punish                         # 责任人  
        self.str_retification_confirmation_number = str_retification_confirmation_number # 整改确认单编号  
        self.str_check_type          = str_check_type                                 # 检查类型 
        
        ########## 数据清洗合并 ##########
        self.retification_requirements_and_deadline =  None # 整改措施与整改期限合并
        self.responsible_person_and_unit            =  None # 责任人/责任单位

        
        # 数据清洗合并执行框
        self.merge_check_str()
        
        
    def function(self):

        wb = Workbook()     # 创建一个 workbook
        
        ws = wb.active      # 获取被激活的 worksheet
        ws.title = "洁源公司自查隐患台账"
        
        ws2 = wb.create_sheet(title="海宜公司整改隐患台账")
        ws3 = wb.create_sheet(title="政府部门整改隐患台账")
        ws4 = wb.create_sheet(title="集团检查整改隐患台账")
        ws5 = wb.create_sheet(title="隐患统计表")
        
        ### 表格格式 ###
        self.align = Alignment( horizontal='center', vertical = 'center', wrap_text = True ) # 居中类
        self.font_normul  = Font( name   = "宋体", size = 12 , bold = True , italic = False , color = "00000000")
        self.font_normul_ = Font( name   = "宋体", size = 12 , bold = False , italic = False , color = "00000000")
        self.font_normul1 = Font( name   = "宋体", size = 11 , bold = False , italic = False , color = "00000000")
        self.font_title   = Font( name   = "宋体", size = 16 , bold = True  , italic = False , color = "00000000")
        self.font_title2   = Font( name   = "宋体", size = 16 , bold = True  , italic = False , color = "fff62100") #红色
        self.border_cell  = Border(left  = Side(style = 'thin', color = '00000000'),
                              right = Side(style = 'thin', color = '00000000'),
                              top   = Side(style = 'thin', color = "00000000"),
                              bottom= Side(style = 'thin', color = "00000000")
                              )
        self.fill_ =  PatternFill(patternType='solid',fgColor='ffe2efda')#淡绿色
        self.fill1 =  PatternFill(patternType='solid',fgColor='92d050')#亮绿色
        self.fill2 =  PatternFill(patternType='solid',fgColor='fdfd30')#亮黄色
        
        ################## 调整表头 ####################
        
        ws['A1'] = "珠海市海宜洁源餐厨垃圾处置有限公司安全生产管理体系"  # 设置单元格内容 
        ws['J1'] = "编号：HY-YH002"
        ws['J2'] = "版本/修订：A/1"
        ws['A3'] = "珠海市海宜洁源餐厨垃圾处置有限公司安全隐患整改台账"
        
        ws2['A1'] = "珠海市海宜环境投资有限公司安全生产管理体系"
        ws2['J1'] = "编号：HY-YH002"
        ws2['J2'] = "版本/修订：A/1"
        ws2['A3'] = "珠海市海宜洁源餐厨垃圾处置有限公司安全隐患整改台账（海宜检查项整改）"
        
        ws3['A1'] = "珠海市海宜环境投资有限公司安全生产管理体系"
        ws3['J1'] = "编号：HY-YH002"
        ws3['J2'] = "版本/修订：A/1"
        ws3['A3'] = "珠海市海宜洁源餐厨垃圾处置有限公司安全隐患整改台账（政府单位检查整改）"
        
        ws4['A1'] = "珠海市海宜环境投资有限公司安全生产管理体系"
        ws4['J1'] = "编号：HY-YH002"
        ws4['J2'] = "版本/修订：A/1"
        ws4['A3'] = "珠海市海宜洁源餐厨垃圾处置有限公司安全隐患整改台账（集团检查整改）"

        ################## 输入内容 ####################
        self.connect_jieyuan_database()           # 加载数据
        
        # 分类
        self.check_haiyi_dict       = self.classify_hz(['海宜查']       ,self.str_check_type,self.str_check_time)        # 检查类型:海宜查
        self.check_jieyuan_dickt    = self.classify_hz(['自查','查收运','查行政与应急'] ,self.str_check_type,self.str_check_time)        # 检查类型:自查
        self.check_jituan_dict      = self.classify_hz(['安保部查']     ,self.str_check_type,self.str_check_time)        # 检查类型:安保部查
        self.check_zfbmc_dict       = self.classify_hz(['政府部门查']   ,self.str_check_type,self.str_check_time)        # 检查类型:政府部门查

        # 生成excel
        self.select_execl_sheet(ws,self.check_jieyuan_dickt) # 洁源的表
        self.select_execl_sheet(ws2,self.check_haiyi_dict)   # 海宜的表
        self.select_execl_sheet(ws3,self.check_zfbmc_dict)   # 政府部门查的表
        self.select_execl_sheet(ws4,self.check_jituan_dict)  # 安保部查的表


        # 统计表sheet
        ws5['A1'] = '以下利用公式自行统计，请勿动本页公式。'
        ws5['A2'] = '洁源餐厨公司自查隐患类型'
        ws5['A3'] = '隐患类型'
        ws5['B3'] = '数量'
        ws5['A4'] = '电气安全隐患'
        ws5['B4'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"电气安全隐患")'
        ws5['A5'] = '管理缺失'
        ws5['B5'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"管理缺失")'
        ws5['A6'] = '火灾安全隐患'
        ws5['B6'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"火灾安全隐患")'
        ws5['A7'] = '人员违反安全管理规定行为'
        ws5['B7'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"人员违反安全管理规定行为")'
        ws5['A8'] = '设备设施的不安全状态'
        ws5['B8'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"设备设施的不安全状态")'
        ws5['A9'] = '危险化学品安全隐患'
        ws5['B9'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"危险化学品安全隐患")'
        ws5['A10'] = '应急管理隐患'
        ws5['B10'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"应急管理隐患")'
        ws5['A11'] = '车辆安全隐患'
        ws5['B11'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"车辆安全隐患")'
        ws5['A12'] = '环保隐患'
        ws5['B12'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"环保隐患")'
        ws5['A13'] = '门岗管理隐患'
        ws5['B13'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"门岗管理隐患")'
        ws5['A14'] = '食品安全隐患'
        ws5['B14'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"食品安全隐患")'
        ws5['A15'] = '八大危险作业管理隐患'
        ws5['B15'] = '=COUNTIF(洁源公司自查隐患台账!D5:D964,"八大危险作业管理隐患")'
        ws5['A16'] = '总计'
        ws5['B16'] = '=SUM(B4:B15)'
        
        ws5['D2'] = '洁源餐厨公司自查隐患级别'
        ws5['D3'] = '隐患级别'
        ws5['E3'] = '数量'
        ws5['D4'] = '一般隐患（班组级）'
        ws5['E4'] = '=COUNTIF(洁源公司自查隐患台账!E5:E964,"一般隐患（班组级）")'
        ws5['D5'] = '一般隐患（厂级）'
        ws5['E5'] = '=COUNTIF(洁源公司自查隐患台账!E5:E964,"一般隐患（厂级）")'
        ws5['D6'] = '一般隐患（公司级）'
        ws5['E6'] = '=COUNTIF(洁源公司自查隐患台账!E5:E964,"一般隐患（公司级）")'
        ws5['D7'] = '总计'
        ws5['E7'] = '=SUM(E4:E6)'
        
        ws5['D9'] = '洁源餐厨公司自查整改统计'
        ws5['D10'] = '已整改'
        ws5['E10'] = '=COUNTIF(洁源公司自查隐患台账!G5:G966,"已完成整改")'
        ws5['D11'] = '未整改'
        ws5['E11'] = '=COUNTIF(洁源公司自查隐患台账!B5:B966,"<>")-E10'
        
        # 调整列宽
        abc = ['A','B','C','D','E']
        row_width = [37.51,20.63,8.38,37.51,37.51]
        for i,linename in enumerate(abc):
            ws5.column_dimensions[abc[i]].width = row_width[i]

        # 调整行高
        ws5.row_dimensions[1].height     = 30             # 调整行高

        # 调整字体及表格样式
        ws5['A1'].font      = self.font_title2
        ws5['A1'].fill      = self.fill2
        ws5['A1'].alignment = self.align                  # 设置居中
        ws5['A1'].border    = self.border_cell            # 设置边框

        for row in ws5['A2:B16']:
            for cell in row:
                cell.alignment = self.align               # 设置居中
                cell.font      = self.font_normul1        # 设置字体
                cell.border    = self.border_cell         # 设置边框

        for row in ws5['D2:E7']:
            for cell in row:
                cell.alignment = self.align               # 设置居中
                cell.font      = self.font_normul1        # 设置字体
                cell.border    = self.border_cell         # 设置边框

        for row in ws5['D9:E11']:
            for cell in row:
                cell.alignment = self.align               # 设置居中
                cell.font      = self.font_normul1        # 设置字体
                cell.border    = self.border_cell         # 设置边框

        ws5['A2'].fill = self.fill1
        ws5['A3'].fill = self.fill1
        ws5['B3'].fill = self.fill1
        ws5['D2'].fill = self.fill1
        ws5['D3'].fill = self.fill1
        ws5['E3'].fill = self.fill1
        ws5['D9'].fill = self.fill1
        
        # 合并单元格
        ws5.merge_cells(range_string = 'A1:E1')
        ws5.merge_cells(range_string = 'A2:B2')
        ws5.merge_cells(range_string = 'D2:E2')
        ws5.merge_cells(range_string = 'D9:E9')
        
        ###############
        # 添加图表
        pie = PieChart()
        labels = Reference(ws5, min_col = 1, min_row = 4, max_row = 15)
        data   = Reference(ws5, min_col = 2, min_row = 3, max_row = 15)
        pie.add_data(data, titles_from_data=True)
        pie.set_categories(labels)
        pie.title = "洁源餐厨公司自查隐患类型"
        
        # 调整图表样式
        pie.legend.position = 'tr'
        pie.legend.layout = Layout( # 图例格式
                manualLayout = ManualLayout(
                                            yMode = 'factor',
                                            xMode = 'factor',
                                            x = -0.1, y = 0,
                                            h = 0.8, w = 0.4
                                                )
                                            )
        
        pie.layout = Layout(        # 图形格式
                manualLayout = ManualLayout(
                                            yMode = 'factor',
                                            xMode = 'factor',
                                            x = 0.03, y = 0,
                                            h = 0.8, w = 0.45
                                                )
                                            )
        # 图表百分比调整
        pie.style  = 2
        s1 = pie.series[0]
        s1.graphicalProperties.line.solidFill = 'fcfcfc'    # 图表边缘线条颜色
        s1.graphicalProperties.line.width = 15000           # 图表边缘线条宽度
        s1.dLbls = DataLabelList()
        s1.dLbls.showPercent = True   # 百分比显示
        s1.dLbls.dLblPos = 'outEnd'   # 标签显示模式：外显  {‘outEnd’, ‘inBase’, ‘b’, ‘inEnd’, ‘l’, ‘ctr’, ‘t’, ‘bestFit’, ‘r’}

        # 调整图表大小
        pie.height = 12
        pie.width  = 13
        
        
        ###############
        # 添加图表
        pie1 = PieChart()
        labels1 = Reference(ws5, min_col = 4, min_row = 4, max_row = 6)
        data1   = Reference(ws5, min_col = 5, min_row = 3, max_row = 6)
        pie1.add_data(data1, titles_from_data=True)
        pie1.set_categories(labels1)
        pie1.title = "洁源餐厨公司自查隐患级别"
        
        # 调整图表样式
        pie1.legend.position = 'tr'
        pie1.legend.layout = Layout( # 图例格式
                manualLayout = ManualLayout(
                                            yMode = 'factor',
                                            xMode = 'factor',
                                            x = -0.1, y = 0.26,
                                            h = 0.4, w = 0.4
                                                )
                                            )
        
        pie1.layout = Layout(        # 图形格式
                manualLayout = ManualLayout(
                                            yMode = 'factor',
                                            xMode = 'factor',
                                            x = 0.1, y = 0,
                                            h = 0.8, w = 0.45
                                                )
                                            )
        # 图表百分比调整
        pie.style  = 2
        s2 = pie1.series[0]
        s2.graphicalProperties.line.solidFill = 'fcfcfc'    # 图表边缘线条颜色
        s2.graphicalProperties.line.width = 15000           # 图表边缘线条宽度
        s2.dLbls = DataLabelList()
        s2.dLbls.showPercent = True   # 百分比显示
        s2.dLbls.dLblPos = 'outEnd'   # 标签显示模式：外显  {‘outEnd’, ‘inBase’, ‘b’, ‘inEnd’, ‘l’, ‘ctr’, ‘t’, ‘bestFit’, ‘r’}

        # 调整图表大小
        pie1.height = 12
        pie1.width  = 15.5
        
        ###############
        ws5.add_chart(pie, "A19")
        ws5.add_chart(pie1,"D19")
        try:
            output_dir = r"{}".format(self.output_dir)
            self.output = os.path.join(output_dir, "洁源公司隐患总台账%s.xlsx"%now5)
            wb.save(self.output)  # 保存 Excel 文件
        except: 
            self.show_warning_message('请关闭当前生成好的隐患总台账excel表')

    def output_file_path(self):
        return self.output

    def select_execl_sheet(self,sheet,select):

        sheet.merge_cells(range_string = 'A1:I2')                       # 合并单元格
        sheet.merge_cells(range_string = 'A3:J3')                       # 合并单元格
        
        for row in sheet['A1:K3']:
            for cell in row:
                cell.alignment = self.align                           # 设置居中
                cell.font      = self.font_title                      # 设置字体
                cell.border    = self.border_cell                     # 设置边框

        line = ['序号','隐患内容','整改措施','隐患类型','隐患级别','检查时间','整改完成情况','完成时间','责任人/责任单位','整改确认单编号','备注']
        
        count = 0
        for row in sheet['A4:K4']:
           for cell in row:
               cell.value = line[count]
               count += 1
               cell.alignment = self.align                           # 设置居中
               cell.font      = self.font_normul                     # 设置字体
               cell.border    = self.border_cell                     # 设置边框

        # 调整列宽
        abc = ['A','B','C','D','E','F','G','H','I','J','K']
        row_width = [8.05,50.01,50.01,27.72,21.38,19.37,15.38,15.7,50.01,21.15,15.09]
        for i,linename in enumerate(abc):
            sheet.column_dimensions[abc[i]].width = row_width[i]
        
        # 调整行高
        sheet.row_dimensions[1].height     = 51      # 调整行高
        sheet.row_dimensions[2].height     = 51
        sheet.row_dimensions[3].height     = 28.5
        sheet.row_dimensions[4].height     = 22.5
        

        ########### 数据输入 ############
        if select:
            start = 1
            for checktime,content in select.items() :
                # print("checktime:%s"%checktime,'content:%s'%content)
                count_start = start + 4
                for point in content:
                    sheet.append([
                        start,
                        self.str_hidden_content[point],# 隐患内容
                        self.retification_requirements_and_deadline[point],# 整改要求与整改期限合并
                        self.str_hidden_type[point],            # 隐患类型
                        self.str_hidden_level[point],           # 隐患级别
                        self.str_check_time[point],             # 检查时间
                        self.str_correactive_situastion[point], # 整改情况
                        self.str_complete_time[point],          # 整改完成时间  
                        self.responsible_person_and_unit[point],# 责任人/责任单位 
                        self.str_retification_confirmation_number[point], # 整改确认单编号  
                        ])
                    start += 1
                    
                    count_end   = count_start + len(content)-1
                    sheet.row_dimensions[start+3].height     = 80              # 调整行高
                # print("count_start:",count_start)
                # print("count_end:",count_end)
                sheet.merge_cells(range_string = 'F%s:F%s'%(count_start,count_end)) # 合并 检查时间 列
                    
            for row in sheet['A5:K%s'%(start+3)]:
                for cell in row:
                    cell.alignment = self.align               # 设置居中
                    cell.font      = self.font_normul_        # 设置字体
                    cell.border    = self.border_cell






    # 预分类器：用于分类输入数据 返回分类好的字典
    def Pre_sorted(self,pre):
        classified_dates_dict  = defaultdict(list)     #
        for index, data in enumerate(pre):
            classified_dates_dict[data].append(index)
        return classified_dates_dict
    
    # 二次分类器：输入要的类型，返回该类型所对应的日期的分类字典，字典里包含有分类好的列表
    def classify_hz(self,object_to_c,object_list,Sec_class_ob):
        '''
        object_to_c：分类目标，以第一个为基准
        object_list：待分类集，此处是检查类型
        str_check_time：二次待分类集，此处是检查时间
        '''
        cache = []
        object_newdict = {}
        sec_newdict    = {}
        cache1 = []
        date = self.Pre_sorted(object_list)
        print(date)
        for data, index_list in date.items():
            #print(data, ":", index_list)
            if data in object_to_c :                        # 选出对象列表
                cache1 += index_list
                cache1.sort( reverse=False)
                object_newdict[object_to_c[0]] = cache1 # 字典
                
        print('第一次分类情况：',object_newdict)
        if object_to_c[0] in object_newdict.keys():
            # 二次分类
            data = [cache.append(Sec_class_ob[i]) for i in object_newdict[object_to_c[0]]]
            # print('第二次预分类:',cache)
            Sec_date = self.Pre_sorted(cache)
            # print('第二次分类情况：',Sec_date )
            for data, index_list in Sec_date.items():
                sec_newdict[data] = [object_newdict[object_to_c[0]][i]  for i in index_list ] # 将列表对应回去
                
            print('最终分类：',sec_newdict)
            
            return sec_newdict
        else:
            return None


    #   数据清洗合并 
    def merge_check_str(self): 
        
        ################# 计算 #################
        retification_requirements_and_deadline = [] # 整改要求与整改期限合并
        responsible_person_and_unit            = [] # 责任人/责任单位

        for i in range(len(self.str_hidden_content)):
            str_right_deadlint = ''
            
            # 整改要求与整改期限合并
            if self.str_corrective_deadline[i] != '':

                if '即时整改' in self.str_corrective_deadline[i] :
                    ...
                elif '-' in self.str_corrective_deadline[i] :
                    str_right_deadlint = '前整改完成'
                    self.str_corrective_deadline[i] = self.str_corrective_deadline[i].replace('-', '年', 1).replace('-', '月').replace('年0','年').replace('月0','月')+'日' # str.replace(old, new[, max])
            else:
                self.str_corrective_deadline[i] = '警告！请填写整改期限' 
                #self.show_warning_message('警告！请填写整改期限') #TODO

            retification_requirements_and_deadline.append(self.str_corrective_measures[i] + '\n（整改期限：%s%s）'%(self.str_corrective_deadline[i],str_right_deadlint  ))

            # 责任人员
            if self.str_responsible_punish[i]!= '':
                if '公司' in  self.str_responsible_punish[i]:
                    str_front = '责任单位：'
                elif ('部' or '中心' ) in  self.str_responsible_punish[i]:
                    str_front = '责任部门：'
                else:
                    str_front = '责任人员：'
            else:
                str_front = ''

            person =  self.str_responsible_punish[i].translate(str.maketrans('-', ' ','0123456789')) # translate(str.maketrans: 第一个和第二个参数的长度必须匹配。在两个参数的情况下，会将第一个参数的字符，依次的映射成第二个参数的字符（o-> X，w-> Y）。第三个参数表示在映射完的结果之后，需要移除的字符。
            responsible_person_and_unit.append( str_front + person )

        # 合并结果
        self.retification_requirements_and_deadline =  retification_requirements_and_deadline # 整改要求与整改期限合并
        self.responsible_person_and_unit            =  responsible_person_and_unit            # 责任人/责任单位

    #%%    显示错误提示信息
    def show_warning_message(self,show_error_str):
        message_box = QMessageBox()
        message_box.setIcon(QMessageBox.Warning)
        message_box.setText("{}".format(show_error_str))
        message_box.setWindowTitle("警告")
        message_box.exec_()




# =============================================================================
#           测试函数
# =============================================================================

'''
        字体大小对应pt
        初号44pt
        小初36pt
        号26pt
        小一24pt
        二号22pt
        小二18pt
        三号16pt
        小三15pt
        四号14pt
        小四12pt
        五号10.5pt
        八号5
        七号5.5
        字号‘小六’对应磅值6.5
        字号‘六号’对应磅值7.5
        字号‘小五’对应磅值9
        '''
    # 统计段落数和行数
def tongji_d_h(file,name):
    num_paragraphs = len(file.paragraphs)
    num_lines = 0
    for paragraph in file.paragraphs:
        num_lines += len(paragraph.text.split('\n'))
        # 统计表格数
    num_tables = len(file.tables)
    for table in file.tables:    
        row_count = len(table.rows)
        column_count = len(table.columns)
        print(f'{name}模板的第{table}个表格中有{row_count}行，{column_count}列')
    print(f"{name}模板中有 {num_paragraphs} 个段落和 {num_lines} 行文本, {num_tables} 个表格。")



# =============================================================================
#           测试
# =============================================================================

# select_data = [0,1,2,3,4,5,6,7,8]
# taxt = overall_hidden_sheet("../main",select_data)
# taxt.function()
